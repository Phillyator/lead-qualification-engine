"""Output writer — writes scored results to xlsx."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))

TIER_FILLS = {
    "A — Hot": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "B — Warm": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "C — Cool": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
    "Disqualified": PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"),
}

OUT_HEADERS = [
    "Rank", "Company Name", "Industry", "Industry Tier",
    "# Employees", "Size Score", "Industry Score", "Keyword Score",
    "Total Score", "Tier", "Keyword Signals", "Flags",
    "Disqualify Reason", "Website", "LinkedIn URL", "Short Description",
]

COL_WIDTHS = {
    1: 6, 2: 40, 3: 30, 4: 8, 5: 12, 6: 10, 7: 12, 8: 12,
    9: 10, 10: 14, 11: 45, 12: 25, 13: 35, 14: 30, 15: 45, 16: 60,
}


def write_results(workbook_path: str, results: list[dict], config: dict):
    """Write scored results to a new tab in the workbook."""
    wb = openpyxl.load_workbook(workbook_path)
    sheet_name = config["output"]["sheet_name"]

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    out = wb.create_sheet(sheet_name)

    # Headers
    for col, header in enumerate(OUT_HEADERS, 1):
        cell = out.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Data
    rank = 0
    for i, r in enumerate(results):
        row_num = i + 2
        if r["tier"] != "Disqualified":
            rank += 1
            display_rank = rank
        else:
            display_rank = "—"

        values = [
            display_rank,
            r["name"],
            r["industry"],
            r["industry_tier"],
            r["employees_num"],
            r["size_score"],
            r["industry_score"],
            r["keyword_score"],
            r["total_score"],
            r["tier"],
            ", ".join(r["keyword_signals"]) if r["keyword_signals"] else "—",
            ", ".join(r["flags"]) if r["flags"] else "—",
            r["disqualify_reason"] or "—",
            r["website"],
            r["linkedin"],
            (r["description"] or "")[:200],
        ]

        fill = TIER_FILLS.get(r["tier"])
        for col, val in enumerate(values, 1):
            cell = out.cell(row=row_num, column=col, value=val)
            cell.border = THIN_BORDER
            if fill and col in (9, 10, 12):
                cell.fill = fill

    # Column widths
    for col, width in COL_WIDTHS.items():
        out.column_dimensions[get_column_letter(col)].width = width

    out.freeze_panes = "A2"
    wb.save(workbook_path)


def print_summary(results: list[dict]):
    """Print scoring summary to stdout."""
    tiers = {}
    for r in results:
        tiers[r["tier"]] = tiers.get(r["tier"], 0) + 1

    print(f"\nScored {len(results)} companies:")
    for tier in ["A — Hot", "B — Warm", "C — Cool", "Disqualified"]:
        print(f"  {tier}: {tiers.get(tier, 0)}")

    print(f"\nTop 20:")
    for r in results[:20]:
        signals = ", ".join(r["keyword_signals"]) if r["keyword_signals"] else "none"
        flag_str = f" ⚠ {', '.join(r['flags'])}" if r["flags"] else ""
        print(f"  {r['total_score']:3d} | {r['tier']:10s} | {r['name']:45s} | {signals}{flag_str}")

    flagged = [r for r in results if r["flags"]]
    if flagged:
        print(f"\nFlagged companies ({len(flagged)}):")
        for r in flagged:
            print(f"  {r['total_score']:3d} | {r['name']:45s} | {', '.join(r['flags'])}")
