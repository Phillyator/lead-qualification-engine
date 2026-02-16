"""Pass 1 scoring engine — scores companies from spreadsheet data against ICP config."""

import re
import openpyxl
from .signals import scan_keywords


def score_size(employees: int | None, config: dict) -> int | None:
    """Score company by employee count. Returns None if disqualified."""
    min_emp = config["size_scoring"].get("min_employees", 0)
    if employees is None or employees < min_emp:
        return None

    for bracket in config["size_scoring"]["brackets"]:
        if employees <= bracket["max"]:
            return bracket["points"]
    return None


def score_industry(industry: str, config: dict) -> tuple[str, int | None]:
    """Score company by industry tier. Returns (tier_letter, points) or (tier, None) if disqualified."""
    tiers = config["industry_tiers"]
    for tier_name in ["A", "B", "C", "OUT"]:
        tier = tiers.get(tier_name, {})
        if industry in tier.get("industries", []):
            points = tier.get("points")
            return tier_name, points
    # Default
    default_tier = tiers.get("default_tier", "C")
    default_points = tiers.get("default_points", 10)
    return default_tier, default_points


def assign_tier(score: int, config: dict) -> str:
    """Assign a tier label based on score thresholds."""
    for tier in config["tiers"]:
        if score >= tier["min_score"]:
            return tier["name"]
    return "Disqualified"


def parse_employees(value) -> int | None:
    """Parse employee count from various formats."""
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str):
        digits = re.sub(r"[^\d]", "", value)
        return int(digits) if digits else None
    return None


def read_companies(workbook_path: str, config: dict) -> list[dict]:
    """Read companies from the source spreadsheet."""
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb[config["input"]["sheet_name"]]
    cols = config["input"]["columns"]
    data_start = config["input"].get("data_start_row", 3)

    companies = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        name = row[cols["company_name"]] if len(row) > cols["company_name"] else None
        if not name:
            continue
        companies.append({
            "name": row[cols["company_name"]],
            "industry": row[cols["industry"]] if len(row) > cols["industry"] else None,
            "employees": row[cols["employees"]] if len(row) > cols["employees"] else None,
            "website": row[cols["website"]] if len(row) > cols["website"] else None,
            "linkedin": row[cols["linkedin"]] if len(row) > cols["linkedin"] else None,
            "revenue": row[cols["revenue"]] if len(row) > cols["revenue"] else None,
            "founded": row[cols["founded"]] if len(row) > cols["founded"] else None,
            "description": row[cols["description"]] if len(row) > cols["description"] else None,
            "keywords": row[cols["keywords"]] if len(row) > cols["keywords"] else None,
        })
    wb.close()
    return companies


def score_companies(companies: list[dict], config: dict) -> list[dict]:
    """Score all companies and return sorted results."""
    results = []

    for c in companies:
        employees = parse_employees(c["employees"])
        size_pts = score_size(employees, config)
        industry_tier, industry_pts = score_industry(c["industry"] or "", config)

        # Check disqualifiers
        disqualify_reasons = []
        if size_pts is None:
            disqualify_reasons.append(f"Too small ({employees or 'unknown'} employees)")
        if industry_pts is None:
            disqualify_reasons.append(f"Industry excluded ({c['industry']})")

        if disqualify_reasons:
            results.append({
                **c,
                "employees_num": employees,
                "size_score": 0,
                "industry_tier": industry_tier,
                "industry_score": 0,
                "keyword_score": 0,
                "keyword_signals": [],
                "flags": [],
                "total_score": 0,
                "tier": "Disqualified",
                "disqualify_reason": "; ".join(disqualify_reasons),
            })
            continue

        # Build text for keyword scanning
        text = f"{c['name'] or ''} {c['description'] or ''} {c['keywords'] or ''}".lower()
        keyword_score, keyword_signals, flags = scan_keywords(text, c["industry"] or "", config)

        total = size_pts + industry_pts + keyword_score
        tier = assign_tier(total, config)

        results.append({
            **c,
            "employees_num": employees,
            "size_score": size_pts,
            "industry_tier": industry_tier,
            "industry_score": industry_pts,
            "keyword_score": keyword_score,
            "keyword_signals": keyword_signals,
            "flags": flags,
            "total_score": total,
            "tier": tier,
            "disqualify_reason": "",
        })

    results.sort(key=lambda x: x["total_score"], reverse=True)
    return results


def read_pass1_results(workbook_path: str, config: dict) -> list[dict]:
    """Read Pass 1 scored results from the output sheet."""
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet_name = config["output"]["sheet_name"]
    ws = wb[sheet_name]

    # Read header row to map column names to indices
    headers = [cell.value for cell in ws[1]]
    col_map = {h: i for i, h in enumerate(headers) if h}

    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[col_map["Company Name"]] if "Company Name" in col_map else None
        if not name:
            continue

        tier = row[col_map.get("Tier", 9)] if "Tier" in col_map else ""
        total_score = row[col_map.get("Total Score", 8)] if "Total Score" in col_map else 0

        results.append({
            "name": name,
            "industry": row[col_map.get("Industry", 2)] or "",
            "industry_tier": row[col_map.get("Industry Tier", 3)] or "",
            "employees_num": row[col_map.get("# Employees", 4)],
            "size_score": row[col_map.get("Size Score", 5)] or 0,
            "industry_score": row[col_map.get("Industry Score", 6)] or 0,
            "keyword_score": row[col_map.get("Keyword Score", 7)] or 0,
            "total_score": total_score or 0,
            "tier": tier or "",
            "keyword_signals": (row[col_map.get("Keyword Signals", 10)] or "").split(", ") if row[col_map.get("Keyword Signals", 10)] and row[col_map.get("Keyword Signals", 10)] != "—" else [],
            "flags": (row[col_map.get("Flags", 11)] or "").split(", ") if row[col_map.get("Flags", 11)] and row[col_map.get("Flags", 11)] != "—" else [],
            "disqualify_reason": row[col_map.get("Disqualify Reason", 12)] if "Disqualify Reason" in col_map else "",
            "website": row[col_map.get("Website", 13)] or "",
            "linkedin": row[col_map.get("LinkedIn URL", 14)] or "",
            "description": row[col_map.get("Short Description", 15)] or "",
        })

    wb.close()
    return results
